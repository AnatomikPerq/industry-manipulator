#!/usr/bin/env python3
"""
Базовый скрипт-парсер СПЕЦИФИКАЦИИ оборудования (СО) - xlsx.

Спецификация - единственный документ связки, который приходит не PDF'ом, а
книгой Excel. Формат - таблица по ГОСТ 21.110 (форма 7):

    строка 1  : шапка ("Позиция", "Наименование и техническая характеристика",
                 "Тип, марка, обозначение документа", "Код оборудования",
                 "Завод-изготовитель", "Единица измерения", "Количество",
                 "Масса единицы", "Примечание" [, коммерческие колонки])
    строка 2  : номера колонок 1..9 - служебная, пропускается
    строка 3  : обозначение документа ("026.809.01.01-ИПК  ША1 СО_25.03.26")
    строки 4+ : позиции спецификации
    в конце   : легенда заливок (" - заказано/оплачено", " - возможна замена", ...)

Колонки НЕ прибиты по буквам: у ША1 их 16 (есть цены), у ШУ-ТМ - 9. Поэтому
шапка ищется по тексту заголовков, а не по индексу - иначе на любом другом
бюро парсер молча поехал бы по колонкам.

ГЛАВНОЕ, что этот парсер обязан сделать правильно - РАЗВЕРНУТЬ колонку
"Позиция" в список отдельных позиционных обозначений. Именно по ним
спецификация сверяется со сборочным чертежом и принципиальной схемой
(bundle_rules.py). В живых файлах встречаются все эти формы записи:

    "A0"                        -> [A0]
    "AI1 AI2"                   -> [AI1, AI2]              (разделитель - пробел)
    "QF01, QF02"                -> [QF01, QF02]            (разделитель - запятая)
    "1KL1...1KL3"               -> [1KL1, 1KL2, 1KL3]      (диапазон, хвостовой номер)
    "TA1 … TA6"                 -> [TA1 ... TA6]           (многоточие-символ U+2026)
    "SF-EL1 …. SF-EL4"          -> [SF-EL1 ... SF-EL4]     (четыре точки - опечатка бюро)
    "1KL1 ... 50KL1"            -> [1KL1 ... 50KL1]        (диапазон по ВЕДУЩЕМУ номеру!)
    "2X-AC ... 4X-AC"           -> [2X-AC, 3X-AC, 4X-AC]   (ведущий номер + суффикс)
    "KL.R, 2K1 … 2K6"           -> [KL.R, 2K1 ... 2K6]     (точка внутри обозначения)

Разворот диапазона сделан не regex'ом "префикс+число", а сравнением ДВУХ концов:
у обоих концов берутся все числовые поля, и если концы отличаются ровно одним
таким полем - разворачиваем по нему. Это единственный способ покрыть и
"1KL1...1KL3" (меняется хвост), и "1KL1 ... 50KL1" (меняется голова) одним
правилом, не гадая, где в обозначении "номер элемента".

Функция extract_to_dir(path, out_dir) -> (files, stats) - контракт ingest.py.
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


# ============================================================
# Шапка таблицы
# ============================================================

# Каноническое имя поля -> подстроки, по которым узнаём заголовок колонки.
# Сравнение по НОРМАЛИЗОВАННОМУ заголовку (без переносов строк и лишних пробелов,
# в нижнем регистре): в живых файлах заголовки набраны с "\n" внутри
# ("Тип, марка,\nобозначение документа\n№ опросного листа").
COLUMN_PATTERNS = [
    # «поз.» и «кол.» (с точкой) - сокращения бюро АТХ: там шапка набрана
    # "Поз." / "Кол." вместо полных слов. Точка в образце обязательна: голое
    # "кол" совпало бы с "количество кабелей" кабельного журнала, а голое
    # "поз" - с любым словом, где эти буквы встретились.
    ("position",     ["позиция", "поз."]),
    ("name",         ["наименование и техническая характеристика", "наименование"]),
    ("type_mark",    ["тип, марка", "тип,марка"]),
    ("code",         ["код оборудования", "код изделия", "код обору"]),
    ("manufacturer", ["завод-изготовитель", "завод изготовитель", "изготовитель"]),
    ("unit",         ["единица измерения", "ед. изм", "единица"]),
    ("quantity",     ["количество", "кол-во", "кол."]),
    ("mass",         ["масса единицы", "масса ед", "масса"]),
    ("note",         ["примечание"]),
]

# Ошибки формул Excel. Попадают в ячейку как текст, когда ссылка уехала
# (удалили строку/лист, перенесли файл). В спецификации это означает, что
# число в этой ячейке потеряно - его нельзя ни прочитать, ни проверить.
EXCEL_ERROR_RE = re.compile(r"^#(REF|VALUE|DIV/0|N/A|NAME\?|NUM|NULL)!?$", re.I)


def _norm_header(v):
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""


def _find_header_row(ws, max_scan=10):
    """Строка шапки = первая строка, где нашлась колонка «Позиция» И
    колонка «Количество». Возвращает (номер строки, {поле: индекс колонки})."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(row=r, column=c).value)
            if not h:
                continue
            for field, patterns in COLUMN_PATTERNS:
                if field in mapping:
                    continue
                if any(h.startswith(p) or p in h for p in patterns):
                    mapping[field] = c
                    break
        if "position" in mapping and "quantity" in mapping:
            return r, mapping
    return None, {}


# ============================================================
# Разбор колонки "Позиция"
# ============================================================

# Разделитель диапазона: "...", "…", "….". Дефис сюда НЕ включён сознательно -
# он встречается ВНУТРИ обозначений ("SF-EL1", "XT-G1", "2X-AC"), и трактовка
# его как диапазона рвала бы обычные обозначения пополам.
#
# Пробелы вокруг разделителя ЗАХВАТЫВАЮТСЯ (\s*): "TA1 … TA6" и "1KL1...1KL3" -
# это одна и та же конструкция, записанная по-разному, и после схлопывания
# пробелов диапазон становится ОДНИМ токеном. Без этого захвата ячейка
# "1KL1...1KL3 2KL1...2KL5 ..." (несколько диапазонов через пробел) разрезалась
# по каждому "..." и давала мусорные "обозначения" вида '1KL3 2KL1' - куски двух
# соседних диапазонов, склеенные пробелом.
# Разделитель диапазона: "1KL1...1KL3", "1KL1…1KL3" и "1K1 - 1K24" (бюро АТХ
# пишет диапазон через дефис). Дефис считается диапазоном ТОЛЬКО с пробелами по
# обе стороны: внутри обозначений он законен ('XT1-G1', 'CB-1L'), и без этого
# требования 'XT1-G1' развалился бы на два конца несуществующего диапазона.
RANGE_SEP_RE = re.compile(r"\s*(?:\.{2,}|…\.*)\s*|\s+[-–—]\s+")

# Служебный маркер: им заменяется разделитель диапазона вместе с пробелами,
# чтобы концы диапазона склеились в один токен и пережили разрез по пробелам.
# \x00 в тексте спецификации встретиться не может.
RANGE_MARK = "\x00"

# Разделители перечисления: запятая, точка с запятой, пробел. Пробел - тоже
# разделитель, потому что ША1 пишет позиции через пробел ("AI1 AI2").
LIST_SEP_RE = re.compile(r"[,;\s]+")

# Числовое поле внутри обозначения.
NUM_FIELD_RE = re.compile(r"\d+")

MAX_RANGE_EXPANSION = 200   # предохранитель от "X1 ... X100000" (опечатка в файле)


def _shape(token):
    """Разбивает обозначение на чередование нечисловых кусков и чисел.
    '1KL1' -> (('', 'KL', ''), (1, 1)); '2X-AC' -> (('', 'X-AC'), (2,))"""
    nums = [int(m.group()) for m in NUM_FIELD_RE.finditer(token)]
    parts = NUM_FIELD_RE.split(token)
    return tuple(parts), tuple(nums)


def _expand_range(left, right):
    """Разворачивает 'left ... right'. Возвращает список обозначений либо None,
    если концы несопоставимы (тогда вызывающий оставит их как есть - две
    отдельные позиции, а не выдуманный диапазон)."""
    lp, ln = _shape(left)
    rp, rn = _shape(right)
    if lp != rp or len(ln) != len(rn) or not ln:
        return None

    diff = [i for i, (a, b) in enumerate(zip(ln, rn)) if a != b]
    if len(diff) != 1:
        return None                      # отличаются нулём или >1 числом - не диапазон
    i = diff[0]
    start, end = ln[i], rn[i]
    if end < start or end - start + 1 > MAX_RANGE_EXPANSION:
        return None

    out = []
    for v in range(start, end + 1):
        nums = list(ln)
        nums[i] = v
        # собираем обратно: parts[0] + num[0] + parts[1] + num[1] + ...
        s = lp[0]
        for j, n in enumerate(nums):
            s += str(n) + (lp[j + 1] if j + 1 < len(lp) else "")
        out.append(s)
    return out


def parse_designators(raw):
    """Колонка «Позиция» -> список отдельных позиционных обозначений.

    Диапазоны разворачиваются, порядок сохраняется, дубли внутри одной ячейки
    убираются (сохраняя первое вхождение).
    """
    if raw is None:
        return []
    text = str(raw).strip()
    if not text or EXCEL_ERROR_RE.match(text):
        return []

    # 1) диапазон вместе с пробелами вокруг него -> один токен через маркер:
    #    "1KL1...1KL3 2KL1...2KL5" -> "1KL1<M>1KL3 2KL1<M>2KL5"
    marked = RANGE_SEP_RE.sub(RANGE_MARK, text)

    # 2) режем перечисление по запятым/пробелам - диапазоны уже неразрывны
    out = []
    for token in LIST_SEP_RE.split(marked):
        token = token.strip()
        if not token:
            continue
        if RANGE_MARK not in token:
            out.append(token)
            continue

        ends = [e.strip() for e in token.split(RANGE_MARK) if e.strip()]
        if len(ends) != 2:
            out.extend(ends)          # "A...B...C" - не гадаем, берём как есть
            continue
        expanded = _expand_range(*ends)
        # концы несопоставимы (напр. "1...7" - номера клемм, а не позиции):
        # оставляем оба конца как есть, выдумывать середину нельзя
        out.extend(expanded if expanded is not None else ends)

    seen, uniq = set(), []
    for d in out:
        d = d.strip(" .,;")
        if d and d not in seen:
            seen.add(d)
            uniq.append(d)
    return uniq


# ============================================================
# Разбор строк
# ============================================================

# Строка-заголовок раздела: есть текст наименования, но нет ни кода, ни
# количества ("Дополнение от 07.05.2026"). Такие строки не позиции, а
# разделители - но всё, что идёт ПОСЛЕ них, относится к этому разделу, и это
# важно: раздел "Дополнительно от 27.04.2026" может ЗАМЕНЯТЬ позицию из
# основной части (та же артикульная строка с другим количеством).
SECTION_RE = re.compile(r"^(дополнен|дополнительно|раздел|изменени)", re.I)

# Строка легенды заливок в конце файла (" - возможна замена").
LEGEND_RE = re.compile(r"^\s*[-–—]\s*\S")


def _cell(ws, row, mapping, field):
    c = mapping.get(field)
    return ws.cell(row=row, column=c).value if c else None


def _is_column_numbers_row(ws, row, mapping):
    """Служебная строка под шапкой, где вместо данных стоят номера колонок
    ("1","2","3",... по ГОСТ). Узнаём по тому, что в найденных колонках стоят
    маленькие целые, идущие по возрастанию слева направо."""
    seen = []
    for field, col in sorted(mapping.items(), key=lambda kv: kv[1]):
        v = ws.cell(row=row, column=col).value
        if v is None or str(v).strip() == "":
            continue
        n = _to_number(v)
        if n is None or n != int(n) or not 1 <= n <= 20:
            return False
        seen.append(int(n))
    return len(seen) >= 3 and seen == sorted(seen)


def _to_number(v):
    """Количество -> float. В файлах встречается и '4,5' (русская запятая),
    и 4.5, и '2'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _clean(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def parse_workbook(path):
    """xlsx -> dict со структурой specification.json."""
    if openpyxl is None:
        raise RuntimeError(
            "Не установлен openpyxl - без него спецификацию (.xlsx) не прочитать. "
            "Установите: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header_row, mapping = _find_header_row(ws)
    if header_row is None:
        raise RuntimeError(
            f"В {os.path.basename(path)} не найдена шапка таблицы: нужны колонки "
            "«Позиция» и «Количество». Это спецификация по ГОСТ 21.110?")

    # Обозначение документа: у обоих бюро оно стоит в колонке наименования
    # ПОД шапкой (строка 3). Между шапкой и им лежит служебная строка с
    # номерами колонок ("1","2","3",...) - её пропускаем: без этого поиск
    # обрывался на ней и обозначение всегда получалось null.
    designation = None
    designation_row = None
    name_col = mapping.get("name", 2)
    for r in range(header_row + 1, min(header_row + 5, ws.max_row) + 1):
        v = _clean(ws.cell(row=r, column=name_col).value)
        if not v:
            continue
        if _is_column_numbers_row(ws, r, mapping):
            continue
        # обозначение = длинный текст без количества и без кода оборудования
        if (len(v) > 5
                and _to_number(_cell(ws, r, mapping, "quantity")) is None
                and not _clean(_cell(ws, r, mapping, "code"))):
            designation = v
            designation_row = r
        break

    items = []
    cell_errors = []
    section = None

    for r in range(header_row + 1, ws.max_row + 1):
        pos_raw = _clean(_cell(ws, r, mapping, "position"))
        name = _clean(_cell(ws, r, mapping, "name"))
        qty = _to_number(_cell(ws, r, mapping, "quantity"))
        code = _clean(_cell(ws, r, mapping, "code"))

        # ошибки формул собираем по ВСЕЙ строке, включая коммерческие колонки
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and EXCEL_ERROR_RE.match(str(v).strip()):
                cell_errors.append({
                    "row": r,
                    "cell": ws.cell(row=r, column=c).coordinate,
                    "value": str(v).strip(),
                })

        if designation_row is not None and r == designation_row:
            continue                                   # строка с обозначением документа
        if _is_column_numbers_row(ws, r, mapping):
            continue                                   # служебная строка «1 2 3 ...» по ГОСТ
        if not name and not code and qty is None:
            continue                                   # пустая строка
        if name and qty is None and code is None and not pos_raw:
            if SECTION_RE.match(name):
                section = name
            # строки легенды заливок в конце файла - не позиции
            continue
        if name and LEGEND_RE.match(name) and qty is None:
            continue

        items.append({
            "row": r,
            "section": section,
            "position_raw": pos_raw,
            "designators": parse_designators(pos_raw),
            "name": name,
            "type_mark": _clean(_cell(ws, r, mapping, "type_mark")),
            "code": code,
            "manufacturer": _clean(_cell(ws, r, mapping, "manufacturer")),
            "unit": _clean(_cell(ws, r, mapping, "unit")),
            "quantity": qty,
            "mass": _to_number(_cell(ws, r, mapping, "mass")),
            "note": _clean(_cell(ws, r, mapping, "note")),
        })

    # индекс: обозначение -> строки, где оно встречается
    by_designator = defaultdict(list)
    for it in items:
        for d in it["designators"]:
            by_designator[d].append(it["row"])

    all_designators = sorted(by_designator)
    counts = Counter(it["code"] for it in items if it["code"])

    return {
        "document_metadata": {
            "source_file": os.path.basename(path),
            "sheet_name": ws.title,
            "designation_in_document": designation,
            "header_row": header_row,
            "columns_found": {k: openpyxl.utils.get_column_letter(v)
                              for k, v in sorted(mapping.items())},
            "extra_sheets": [w.title for w in wb.worksheets[1:]],
        },
        "column_legend": {
            "position_raw": "колонка «Позиция» как есть в файле",
            "designators": "то же, развёрнутое в отдельные позиционные обозначения "
                           "(диапазоны '1KL1...1KL3' раскрыты) - ГЛАВНЫЙ КЛЮЧ СВЕРКИ "
                           "со сборочным чертежом и принципиальной схемой",
            "name": "наименование и техническая характеристика",
            "type_mark": "тип, марка, обозначение документа",
            "code": "код оборудования (артикул производителя)",
            "quantity": "количество в единицах измерения (unit)",
            "section": "раздел спецификации, в котором стоит строка (напр. "
                       "«Дополнение от 07.05.2026»); null - основная часть",
        },
        "domain_notes_for_analysis": (
            "1) Количество НЕ обязано совпадать с числом позиционных обозначений: "
            "для клеммников в «Позиция» стоят ОБОЗНАЧЕНИЯ КЛЕММНИКОВ (напр. '2X2 3X2'), "
            "а в «Количество» - число КЛЕММ в них (27). Количество БОЛЬШЕ числа "
            "обозначений - норма. "
            "2) Одно и то же обозначение в нескольких строках - тоже норма: у устройства "
            "есть аксессуары (реле + колодка + фиксатор + шильдик - 4 строки на одно 'KL1'). "
            "3) Строки без «Позиции» (короба, DIN-рейки, крепёж) - расходные материалы, "
            "у них нет позиционного обозначения и сверять их со схемой не с чем. "
            "4) Разделы «Дополнение от ...» могут ЗАМЕНЯТЬ строку основной части: тот же "
            "код оборудования с другим количеством - это ревизия, а не дубль."
        ),
        "items": items,
        "designator_index": {d: by_designator[d] for d in all_designators},
        "cell_errors": cell_errors,
        "statistics": {
            "total_rows": len(items),
            "rows_with_position": sum(1 for i in items if i["designators"]),
            "unique_designators": len(all_designators),
            "total_designators": sum(len(i["designators"]) for i in items),
            "rows_without_code": sum(1 for i in items if not i["code"]),
            "duplicate_codes": {k: v for k, v in counts.items() if v > 1},
            "cell_errors": len(cell_errors),
            "sections": sorted({i["section"] for i in items if i["section"]}),
        },
    }


# ============================================================
# Контракт ingest.py
# ============================================================

def extract_to_dir(xlsx_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    doc = parse_workbook(xlsx_path)

    path = os.path.join(out_dir, "specification.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    st = doc["statistics"]
    stats = {
        "spec_rows": st["total_rows"],
        "spec_unique_designators": st["unique_designators"],
        "spec_cell_errors": st["cell_errors"],
        "spec_designation": doc["document_metadata"]["designation_in_document"],
    }
    return ["specification.json"], stats


def main():
    if len(sys.argv) < 2:
        print("Использование: python3 specification_to_json.py path/to/СО.xlsx [out_dir]")
        sys.exit(1)
    out = sys.argv[2] if len(sys.argv) > 2 else "."
    files, stats = extract_to_dir(sys.argv[1], out)
    print(json.dumps({"files": files, "stats": stats}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
