#!/usr/bin/env python3
"""
Вызов модели для обычного чата: сборка сообщений и ПОТОКОВЫЙ ответ.

Здесь живёт всё, что тянет тяжёлые зависимости (openai, fitz, openpyxl), -
ChatStore же (chats.py) на голой стандартной библиотеке и только хранит диалог.
Импорты этих пакетов - ЛЕНИВЫЕ, внутри функций: тот же приём, что у report_pdf
и fragment в server.py (fitz нужен серверу только в своих обработчиках).

ПОЧЕМУ ТЕКСТ ИЗ ФАЙЛОВ ИЗВЛЕКАЕМ САМИ. Мы говорим с LM Studio через его API
(/v1/chat/completions), а не через его GUI: «прикрепить PDF, и модель сама
разберётся» - это функция графической оболочки, по API такого нет. Чтобы ВСЕ
файлы пользователя дошли до модели (требование заказчика), их содержимое надо
передать явно: картинки - как image-части (vision), остальное - извлечённым
текстом прямо в сообщении. Сервер ИИ может стоять на другой машине (адрес в
config.local.yaml), поэтому картинка уходит data-URL'ом, а не путём к файлу -
файловой системы этой машины сервер не видит (ровно как в llm_client._image_part).
"""

import base64
import logging
import time
from pathlib import Path

logger = logging.getLogger(__name__)

# Разговорная температура: у анализа в config.yaml стоит 0.2 (нужна точность), а
# чат - живой диалог, и небольшая свобода здесь уместнее. Значение своё, а не из
# config.yaml, потому что config.yaml выверен под совсем другую задачу.
CHAT_TEMPERATURE = 0.7
# Длина ответа. Щедро: чат стримится, пользователь видит текст по мере генерации
# и может остановить кнопкой, так что «слишком длинно» ему не грозит.
CHAT_MAX_TOKENS = 8192

# Сколько символов извлечённого из ОДНОГО файла текста передаём модели. Больше
# незачем: контекст у локальных моделей невелик, а весь смысл - дать модели
# понять, о чём файл, а не загрузить в неё роман.
PER_FILE_CHARS = 20000

# Расширения, из которых текст читается как есть (без разбора формата).
TEXT_EXTS = {
    ".txt", ".md", ".markdown", ".csv", ".tsv", ".json", ".yaml", ".yml",
    ".log", ".ini", ".cfg", ".conf", ".xml", ".html", ".htm", ".rtf",
    ".py", ".js", ".ts", ".css", ".c", ".h", ".cpp", ".java", ".go", ".rs",
    ".sh", ".bat", ".sql",
}

_IMAGE_MIME = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
}


def server_cfg_from_config(config: dict) -> dict:
    """Адрес и ключ сервера ИИ для чата - берём у agent_1 (base_url там уже
    слит с config.local.yaml через pipeline.load_config). Модель НЕ отсюда: её
    выбирает пользователь в самом чате."""
    s = config["llm_servers"]["agent_1"]
    return {"base_url": s["base_url"], "api_key": s.get("api_key")}


# ------------------------------------------------------------------
# Извлечение текста из приложенных файлов
# ------------------------------------------------------------------

def _truncate(text: str, limit: int = PER_FILE_CHARS) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n…[текст обрезан, показано {limit} из {len(text)} символов]"


def extract_file_text(path: Path) -> str | None:
    """Текст из файла для передачи модели, либо None, если извлечь нечего
    (двоичный файл неизвестного формата). Ошибку извлечения не роняем наверх -
    возвращаем None: один нечитаемый файл не должен ломать всё сообщение."""
    suffix = path.suffix.lower()
    try:
        if suffix in TEXT_EXTS:
            return _truncate(path.read_text(encoding="utf-8", errors="replace"))
        if suffix == ".pdf":
            return _extract_pdf(path)
        if suffix in (".xlsx", ".xlsm"):
            return _extract_xlsx(path)
        if suffix == ".docx":
            return _extract_docx(path)
    except Exception as e:  # noqa: BLE001 - извлечение ненадёжно по природе
        logger.warning("Не удалось извлечь текст из %s: %s", path.name, e)
        return None
    return None


def _extract_pdf(path: Path) -> str:
    import fitz  # ленивый импорт: тот же приём, что у fragment/report_pdf

    parts = []
    total = 0
    with fitz.open(path) as doc:
        for page in doc:
            parts.append(page.get_text("text"))
            total += len(parts[-1])
            if total > PER_FILE_CHARS:
                break
    return _truncate("\n".join(parts))


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    lines = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            lines.append(f"# Лист: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append("\t".join(cells))
                if sum(len(x) for x in lines) > PER_FILE_CHARS:
                    return _truncate("\n".join(lines))
    finally:
        wb.close()
    return _truncate("\n".join(lines))


def _extract_docx(path: Path) -> str:
    """Грубое извлечение текста из .docx без python-docx: .docx - это zip, а
    текст лежит в word/document.xml. Абзацы - <w:p>, разрывы строк - тегами;
    остальные теги просто выкидываем. Для чата этого достаточно."""
    import re
    import zipfile
    from html import unescape

    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", "replace")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<w:tab\b[^>]*/>", "\t", xml)
    text = re.sub(r"<[^>]+>", "", xml)
    return _truncate(unescape(text))


# ------------------------------------------------------------------
# Сборка сообщений для API
# ------------------------------------------------------------------

def _image_part(path: Path) -> dict:
    mime = _IMAGE_MIME.get(path.suffix.lower(), "image/png")
    data = base64.b64encode(path.read_bytes()).decode()
    return {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{data}"}}


def build_messages(chat: dict, resolve) -> list:
    """Сообщения диалога в формате OpenAI-совместимого API.

    resolve(path) -> Path - преобразователь ссылки файла в абсолютный путь (у
    сервера это ChatStore.resolve_file, который заодно стережёт выход за пределы
    папки чата). Картинки уходят image-частями, прочие файлы - извлечённым
    текстом, дописанным в сообщение.
    """
    messages = []
    for msg in chat.get("messages", []):
        role = msg.get("role")
        text = msg.get("content") or ""
        files = msg.get("files") or []

        if role == "assistant":
            messages.append({"role": "assistant", "content": text})
            continue

        # user. Без файлов - обычная строка (так проще и совместимее); с
        # файлами - список частей (текст + картинки + врезки из файлов).
        if not files:
            messages.append({"role": "user", "content": text})
            continue

        parts = []
        extra_text = []
        for ref in files:
            try:
                path = resolve(ref.get("path"))
            except Exception as e:  # noqa: BLE001 - файл могли удалить руками
                extra_text.append(f"[Файл «{ref.get('name')}» недоступен: {e}]")
                continue
            if ref.get("kind") == "image":
                parts.append(_image_part(path))
            else:
                body = extract_file_text(path)
                if body is None:
                    extra_text.append(
                        f"[Приложен файл «{ref.get('name')}» — извлечь текст не удалось]")
                else:
                    extra_text.append(
                        f"[Содержимое файла «{ref.get('name')}»]\n{body}")

        combined = "\n\n".join([t for t in [text] + extra_text if t])
        # Текст ставим первым куском - модель читает подпись раньше картинки.
        parts.insert(0, {"type": "text", "text": combined or "(без текста)"})
        messages.append({"role": "user", "content": parts})
    return messages


class ThinkSplitter:
    """Разделяет поток на видимый ответ и рассуждение по инлайновым тегам
    <think>…</think>.

    ЗАЧЕМ. Часть «думающих» моделей шлёт рассуждение отдельным полем
    delta.reasoning_content (его ловим прямо в stream_reply), но часть
    ВСТАВЛЯЕТ его в content тегами <think>…</think>. Тег может прийти разрезанным
    между двумя чанками («<th» + «ink>»), поэтому держим хвост, который ещё может
    оказаться началом тега, и не отдаём его, пока не станет ясно.
    """

    OPEN = "<think>"
    CLOSE = "</think>"

    def __init__(self):
        self.in_think = False
        self.buf = ""

    def feed(self, text: str):
        """text -> список (kind, text), kind in {"content","reasoning"}."""
        self.buf += text
        out = []
        while True:
            tag = self.CLOSE if self.in_think else self.OPEN
            idx = self.buf.find(tag)
            if idx == -1:
                keep = self._partial_tail(self.buf, tag)
                emit = self.buf[:len(self.buf) - keep] if keep else self.buf
                if emit:
                    out.append(("reasoning" if self.in_think else "content", emit))
                self.buf = self.buf[len(self.buf) - keep:] if keep else ""
                break
            before = self.buf[:idx]
            if before:
                out.append(("reasoning" if self.in_think else "content", before))
            self.in_think = not self.in_think
            self.buf = self.buf[idx + len(tag):]
        return out

    def flush(self):
        """Остаток в конце потока: отдать как есть (недокрытый тег - просто текст)."""
        if not self.buf:
            return []
        out = [("reasoning" if self.in_think else "content", self.buf)]
        self.buf = ""
        return out

    @staticmethod
    def _partial_tail(s: str, tag: str) -> int:
        """Длина хвоста s, который может быть началом tag (иначе 0)."""
        for k in range(min(len(tag) - 1, len(s)), 0, -1):
            if s.endswith(tag[:k]):
                return k
        return 0


# Как часто слать событие со скоростью генерации: чаще нет смысла - глаз столько
# не читает, а браузер лишний раз перерисовывает счётчик.
_STATS_INTERVAL = 0.4


def stream_reply(server_cfg: dict, model: str, messages: list,
                 temperature: float = CHAT_TEMPERATURE,
                 max_tokens: int = CHAT_MAX_TOKENS):
    """Генератор СОБЫТИЙ ответа модели (потоковый chat.completions). Каждое
    событие - словарь одного из видов:

      {"type": "content",   "text": ...}   - кусок видимого ответа;
      {"type": "reasoning", "text": ...}   - кусок рассуждения «думающей» модели
                                             (свёрнутый блок в интерфейсе);
      {"type": "stats", "tokens": N, "seconds": T, "tps": X}  - скорость генерации.

    Рассуждение показываем (в отличие от прежней версии, где его прятали): по
    просьбе заказчика в чате должен быть виден процесс раздумий, как в привычных
    чатах. Источник рассуждения - либо отдельное поле delta.reasoning_content
    (LM Studio), либо инлайновые теги <think> в content (ThinkSplitter).

    Токены считаем по чанкам: llama.cpp/LM Studio стримит по одному токену на
    чанк. Финальную цифру уточняем по usage, если сервер его прислал
    (stream_options.include_usage).
    """
    from openai import OpenAI

    client = OpenAI(base_url=server_cfg["base_url"],
                    api_key=server_cfg.get("api_key") or "not-needed")
    stream = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=temperature,
        max_tokens=max_tokens,
        stream=True,
        stream_options={"include_usage": True},
    )

    splitter = ThinkSplitter()
    t_first = None
    tokens = 0
    last_stat = 0.0
    usage_tokens = None

    for chunk in stream:
        usage = getattr(chunk, "usage", None)
        if usage is not None:
            usage_tokens = getattr(usage, "completion_tokens", None) or usage_tokens
        if not chunk.choices:
            continue
        delta = chunk.choices[0].delta
        # Рассуждение отдельным полем: имя у разных серверов разное.
        reasoning = (getattr(delta, "reasoning_content", None)
                     or getattr(delta, "reasoning", None))
        piece = getattr(delta, "content", None)
        if not reasoning and not piece:
            continue

        now = time.monotonic()
        if t_first is None:
            t_first = now
        tokens += 1                          # один чанк ≈ один токен

        if reasoning:
            yield {"type": "reasoning", "text": reasoning}
        if piece:
            for kind, text in splitter.feed(piece):
                yield {"type": kind, "text": text}

        elapsed = now - t_first
        if elapsed > 0 and now - last_stat >= _STATS_INTERVAL:
            last_stat = now
            yield {"type": "stats", "tokens": tokens, "seconds": round(elapsed, 2),
                   "tps": round(tokens / elapsed, 1)}

    for kind, text in splitter.flush():
        yield {"type": kind, "text": text}

    # Финальная сводка: цифру токенов берём у сервера, если он её прислал.
    if t_first is not None:
        elapsed = max(time.monotonic() - t_first, 1e-6)
        final_tokens = usage_tokens or tokens
        yield {"type": "stats", "tokens": final_tokens, "seconds": round(elapsed, 2),
               "tps": round(final_tokens / elapsed, 1), "final": True}
